import openpyxl
import os
from openpyxl.utils import get_column_letter

def update_excel(file_path):
    if not os.path.exists(file_path):
        print("El archivo no existe. Creando uno nuevo...")
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Nombre", "Nota", "Asignatura"])
        wb.save(file_path)
    else:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

    new_row = ["Juan", "80", "Matemáticas"]
    sheet.append(new_row)

    try:
        wb.save(file_path)
        print("Los datos fueron agregados correctamente al archivo Excel.")
    except PermissionError:
        print("Error: No se puede guardar el archivo. Asegúrate de que esté cerrado.")

    # Ajuste de ancho de las columnas
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    wb.close()

# Verificar la extensión del archivo
file_path = "notas.xlsx"
if file_path.endswith('.xlsx'):
    update_excel(file_path)
else:
    print("Error: El archivo debe tener la extensión '.xlsx'.")
